from datetime import date, datetime, time
from io import BytesIO

from openpyxl import load_workbook


class ExcelExtractionStrategy:
    def _normalize_cell(self, value):
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned or None
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        return value

    def _build_result(self, columns: list[str], row_values: list[list]) -> dict:
        items = []
        for idx, row in enumerate(row_values, start=1):
            item = {}
            for col_index, column_name in enumerate(columns):
                cell = row[col_index] if col_index < len(row) else None
                item[column_name] = self._normalize_cell(cell)
            items.append({"row_number": idx, "item": item})

        return {
            "columns": columns,
            "items": items,
            "row_count": len(items),
        }

    def _extract_xlsx(self, file_bytes: bytes, file_name: str) -> dict:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        sheet_names = workbook.sheetnames

        if len(sheet_names) > 1:
            print(
                f"Warning: Multiple sheets detected in '{file_name}'. "
                f"Using first sheet '{sheet_names[0]}'.",
                flush=True,
            )

        sheet = workbook[sheet_names[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {
                "file_name": file_name,
                "result": {"columns": [], "items": [], "row_count": 0},
                "meta": {
                    "source": "excel-parser",
                    "sheet_name": sheet_names[0],
                    "sheet_count": len(sheet_names),
                    "format": "xlsx",
                },
            }

        header_row = rows[0]
        columns = []
        for idx, raw_header in enumerate(header_row, start=1):
            normalized = self._normalize_cell(raw_header)
            columns.append(str(normalized) if normalized is not None else f"column_{idx}")

        result = self._build_result(columns, rows[1:])

        return {
            "file_name": file_name,
            "result": result,
            "meta": {
                "source": "excel-parser",
                "sheet_name": sheet_names[0],
                "sheet_count": len(sheet_names),
                "format": "xlsx",
            },
        }

    def _extract_xls(self, file_bytes: bytes, file_name: str) -> dict:
        try:
            import pandas as pd
        except ImportError as exc:
            raise RuntimeError(
                "Parsing .xls requires pandas (and typically xlrd) to be installed."
            ) from exc

        workbook = pd.ExcelFile(BytesIO(file_bytes))
        sheet_names = workbook.sheet_names

        if len(sheet_names) > 1:
            print(
                f"Warning: Multiple sheets detected in '{file_name}'. "
                f"Using first sheet '{sheet_names[0]}'.",
                flush=True,
            )

        df = workbook.parse(sheet_name=sheet_names[0], dtype=object)

        columns = [
            str(col).strip() if str(col).strip() else f"column_{i + 1}"
            for i, col in enumerate(df.columns)
        ]
        row_values = df.where(df.notna(), None).values.tolist()
        result = self._build_result(columns, row_values)

        return {
            "file_name": file_name,
            "result": result,
            "meta": {
                "source": "excel-parser",
                "sheet_name": sheet_names[0],
                "sheet_count": len(sheet_names),
                "format": "xls",
            },
        }

    def extract_data(self, file_bytes: bytes, file_name: str) -> dict:
        lower_name = file_name.lower()
        if lower_name.endswith(".xlsx"):
            return self._extract_xlsx(file_bytes, file_name)
        if lower_name.endswith(".xls"):
            return self._extract_xls(file_bytes, file_name)
        raise ValueError(f"Unsupported Excel file extension for '{file_name}'")


def get_excel_extraction_strategy() -> ExcelExtractionStrategy:
    return ExcelExtractionStrategy()
